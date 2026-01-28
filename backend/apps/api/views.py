"""
API Views for LINE LIFF Influencer Platform
"""

import json
import requests
from django.conf import settings
from django.shortcuts import get_object_or_404
from django.http import StreamingHttpResponse, HttpResponseRedirect
from django.db import transaction, connection
from django.db.models import Case, When, Value, IntegerField, Q
from django.utils import timezone
from rest_framework import status, permissions
from rest_framework.views import APIView
from rest_framework.generics import ListAPIView, RetrieveAPIView, CreateAPIView
from rest_framework.parsers import MultiPartParser, FormParser
from rest_framework.response import Response
from rest_framework.permissions import AllowAny, IsAuthenticated
from rest_framework_simplejwt.tokens import RefreshToken
import logging
import threading
from decimal import Decimal
from apps.api.services.gemini import GeminiService

logger = logging.getLogger(__name__)

from apps.audit_logs.utils import log_action

from apps.users.models import User
from apps.influencers.models import Interest, InfluencerProfile, SocialPlatformAccount
from apps.campaigns.models import Campaign, CampaignApplication

from .serializers import (
    InterestSerializer, UserSerializer, UserWithProfileSerializer,
    RegistrationSerializer, CampaignListSerializer, CampaignDetailSerializer,
    ApplicationSerializer, SubmitWorkSerializer, ValidateDriveLinkSerializer,
    SocialPlatformAccountSerializer, SocialConnectSerializer,
    InfluencerApprovalSerializer, CampaignCreateSerializer,
    CampaignUpdateSerializer
)
from .services import GoogleDriveValidator
from apps.influencers.services import SocialPlatformService


class HealthCheckView(APIView):
    """Health check endpoint."""
    
    permission_classes = [AllowAny]
    
    def get(self, request):
        status_data = {
            'status': 'healthy',
            'message': 'Backend is running',
            'info': 'DB check temporarily disabled for debug'
        }
        return Response(status_data)


class LineLoginView(APIView):
    """
    Authenticate user via LINE LIFF.
    
    Accepts LINE ID token, verifies it, and returns JWT auth token.
    In development mode with DEBUG=True, accepts mock tokens.
    """
    
    permission_classes = [AllowAny]
    
    def post(self, request):
        print(f"[LineLogin] LOGIN ATTEMPT: {list(request.data.keys())}")
        try:
            id_token = request.data.get('id_token')
            access_token = request.data.get('access_token')
            
            if not id_token:
                print("[LineLogin] Result: 400 - id_token required")
                return Response(
                    {'error': 'id_token_required', 'message': 'LINE ID token is required'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            # 1. Verification Step
            if settings.DEBUG and str(id_token).startswith('mock_'):
                print(f"[LineLogin] MOCK TOKEN Mapped: {id_token}")
                line_user_id = request.data.get('line_user_id', 'U_DEV_12345')
                display_name = request.data.get('display_name', 'Dev User')
                picture_url = request.data.get('picture_url', '')
            else:
                try:
                    profile = self._verify_line_token(id_token, access_token)
                    # Normalize profile data
                    line_user_id = profile.get('userId') or profile.get('sub')
                    display_name = profile.get('displayName') or profile.get('name', '')
                    picture_url = profile.get('pictureUrl') or profile.get('picture', '')
                    
                    if not line_user_id:
                        print(f"[LineLogin] Error: Missing user identifier in profile: {profile}")
                        raise Exception("Could not find userId or sub in LINE profile response")
                    
                    print(f"[LineLogin] Verified LINE ID: {line_user_id}")
                except Exception as e:
                    print(f"[LineLogin] Result: 401 - Verification Error: {str(e)}")
                    return Response(
                        {'error': 'invalid_token', 'message': f'LINE Verification Error: {str(e)}'},
                        status=status.HTTP_401_UNAUTHORIZED
                    )
            
            # 2. Database/User Step
            try:
                with transaction.atomic():
                    user, created = User.objects.get_or_create(
                        line_user_id=line_user_id,
                        defaults={
                            'username': line_user_id,
                            'display_name': display_name,
                            'picture_url': picture_url,
                            'status': 'NEW'
                        }
                    )
                    
                    if not created:
                        print(f"[LineLogin] Existing user: {user.username}")
                        user.display_name = display_name
                        user.picture_url = picture_url
                        user.save(update_fields=['display_name', 'picture_url', 'updated_at'])
                    else:
                        print(f"[LineLogin] New user created: {user.username}")
                        # Handle username uniqueness if needed (get_or_create handles it via defaults but if collision happens...)
                        # actually User.objects.get_or_create(line_user_id=...) is safest.

            except Exception as e:
                print(f"[LineLogin] Result: 500 - Database Error: {str(e)}")
                import traceback
                print(traceback.format_exc())
                return Response(
                    {'error': 'db_error', 'message': str(e)},
                    status=status.HTTP_500_INTERNAL_SERVER_ERROR
                )
            
            # 3. Final Step: JWT and Response
            try:
                refresh = RefreshToken.for_user(user)
                response_data = {
                    'token': str(refresh.access_token),
                    'refresh': str(refresh),
                    'user': UserSerializer(user).data
                }
                print(f"[LineLogin] Result: 200 - OK for {user.username}")
                return Response(response_data)
            except Exception as e:
                print(f"[LineLogin] Result: 500 - JWT/Serialization Error: {str(e)}")
                raise e

        except Exception as e:
            print(f"[LineLogin] Result: 500 - Catch-all Exception: {str(e)}")
            import traceback
            tb = traceback.format_exc()
            print(tb)
            return Response(
                {'error': 'server_error', 'message': str(e), 'traceback': tb},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    def _verify_line_token(self, id_token, access_token=None):
        import requests
        
        # 1. Try with access token if available
        if access_token:
            print("[LineLogin] Try /v2/profile")
            response = requests.get(
                'https://api.line.me/v2/profile',
                headers={'Authorization': f'Bearer {access_token}'},
                timeout=10
            )
            if response.status_code == 200:
                print("[LineLogin] Profile success")
                return response.json()
            print(f"[LineLogin] /v2/profile failed: {response.status_code}")
        
        # 2. Fallback to ID token verification
        print(f"[LineLogin] Try /oauth2/v2.1/verify - ClientID: {settings.LINE_CHANNEL_ID}")
        response = requests.post(
            'https://api.line.me/oauth2/v2.1/verify',
            data={
                'id_token': id_token,
                'client_id': settings.LINE_CHANNEL_ID
            },
            timeout=10
        )
        
        if response.status_code == 200:
            print("[LineLogin] Verify success")
            return response.json()
        
        print(f"[LineLogin] LINE API Error: {response.status_code} - {response.text}")
        raise Exception(f"LINE API error {response.status_code}: {response.text}")


class CurrentUserView(APIView):
    """Get current authenticated user's information."""
    
    permission_classes = [IsAuthenticated]
    
    def get(self, request):
        serializer = UserWithProfileSerializer(request.user, context={'request': request})
        return Response(serializer.data)


class InterestListView(ListAPIView):
    """List all available interest categories."""
    
    permission_classes = [AllowAny]
    queryset = Interest.objects.filter(is_active=True)
    serializer_class = InterestSerializer


class RegistrationSubmitView(APIView):
    """Submit registration form."""
    
    permission_classes = [IsAuthenticated]
    
    def compress_uploaded_image(self, uploaded_file):
        """Compress image to reduce file size."""
        if not uploaded_file:
            return None
        
        try:
            from PIL import Image
            from io import BytesIO
            from django.core.files.uploadedfile import InMemoryUploadedFile
            import sys
            
            uploaded_file.seek(0)
            image = Image.open(uploaded_file)
            
            # Convert to RGB (in case of PNG with transparency)
            if image.mode in ('RGBA', 'P'):
                image = image.convert('RGB')
                
            # Resize if max dimension > 1600
            max_size = 1600
            if max(image.size) > max_size:
                ratio = max_size / float(max(image.size))
                new_size = (int(image.size[0] * ratio), int(image.size[1] * ratio))
                image = image.resize(new_size, Image.Resampling.LANCZOS)
                
            output = BytesIO()
            # Compress quality=70
            image.save(output, format='JPEG', quality=70, optimize=True)
            output.seek(0)
            
            return InMemoryUploadedFile(
                output,
                'ImageField',
                f"{uploaded_file.name.rsplit('.', 1)[0]}.jpg",
                'image/jpeg',
                output.getbuffer().nbytes,
                None
            )
        except Exception as e:
            logger.error(f"Image compression failed: {e}")
            uploaded_file.seek(0)
            return uploaded_file # Fallback to original
    
    @transaction.atomic
    def post(self, request):
        # Check if user already has a profile
        if hasattr(request.user, 'profile'):
            return Response(
                {'error': 'already_registered', 'message': 'You have already completed registration'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # 1. Pull data into a clean dictionary to avoid QueryDict issues with ListFields
        # We use request.data.dict() if it's a QueryDict, but that collapses lists.
        # So we manually build a dict and handle 'interests' specifically.
        if hasattr(request.data, 'getlist'):
            # It's a QueryDict (multipart/form-data)
            data = request.data.dict()
            # Restore interests as a list if it exists
            interests_raw = request.data.get('interests')
        else:
            # It's a standard dict (application/json)
            data = request.data.copy()
            interests_raw = data.get('interests')
        
        # 2. Robust Interest Normalization
        # Goal: End up with a list of strings (IDs)
        normalized_interests = []
        
        if interests_raw:
            try:
                # Case A: JSON string (most likely from our frontend)
                if isinstance(interests_raw, str) and (interests_raw.startswith('[') or interests_raw.startswith('{')):
                    import json
                    parsed = json.loads(interests_raw)
                    if isinstance(parsed, list):
                        interests_list = parsed
                    else:
                        interests_list = [parsed]
                # Case B: Already a list
                elif isinstance(interests_raw, list):
                    interests_list = interests_raw
                # Case C: Comma-separated string or single value
                elif isinstance(interests_raw, str):
                    interests_list = [i.strip() for i in interests_raw.split(',') if i.strip()]
                else:
                    interests_list = [interests_raw]
                
                # Case D: Extract IDs from objects if necessary
                for item in interests_list:
                    if item is None:
                        continue
                    if isinstance(item, dict) and 'id' in item:
                        normalized_interests.append(str(item['id']))
                    elif isinstance(item, dict):
                        # Use first key as ID if no 'id' field? Unlikely but safe
                        pass 
                    else:
                        normalized_interests.append(str(item))
            except Exception as e:
                print(f"[Register] Error normalizing interests: {e}")
                # Fallback: try to use whatever was there
                pass

        data['interests'] = normalized_interests
        
        # 3. Handle file uploads with Compression
        if 'id_card_front' in request.FILES:
            data['id_card_front'] = self.compress_uploaded_image(request.FILES['id_card_front'])
        if 'bank_book' in request.FILES:
            data['bank_book'] = self.compress_uploaded_image(request.FILES['bank_book'])

        serializer = RegistrationSerializer(data=data)
        if not serializer.is_valid():
            print(f"[Register] Validation failed: {serializer.errors}")
            return Response(
                {'success': False, 'errors': serializer.errors},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        data = serializer.validated_data
        
        # Create profile
        profile = InfluencerProfile.objects.create(
            user=request.user,
            full_name_th=data['full_name_th'],
            phone=data['phone'],
            email=data.get('email', ''),
            date_of_birth=data['date_of_birth'],
            house_no=data['house_no'],
            village=data.get('village', ''),
            moo=data.get('moo', ''),
            soi=data.get('soi', ''),
            road=data.get('road', ''),
            sub_district=data['sub_district'],
            district=data['district'],
            province=data['province'],
            zipcode=data['zipcode'],
            allow_boost=data.get('allow_boost', False),
            boost_price=data.get('boost_price'),
            allow_original_file=data.get('allow_original_file', False),
            original_file_price=data.get('original_file_price'),
            accept_gifted_video=data.get('accept_gifted_video', False),
            accept_affiliate=data.get('accept_affiliate', False),
            id_card_front=data['id_card_front'],
            bank_book=data['bank_book'],
        )
        
        # Add interests
        interest_ids = data.get('interests', [])
        print(f"[Register] Final interest IDs to save (validated): {interest_ids}")
        
        # Diagnostic: List all interest IDs in DB
        all_db_ids = list(Interest.objects.values_list('id', flat=True))
        print(f"[Register] Available Interest IDs in DB: {all_db_ids}")
        
        interests_qs = Interest.objects.filter(id__in=interest_ids)
        print(f"[Register] Found {interests_qs.count()} match(es) in database for submitted IDs")
        
        for inst in interests_qs:
            print(f"[Register] - Matched: {inst.id}")
            
        profile.interests.set(interests_qs)
        print(f"[Register] Successfully set interests on profile ID: {profile.id}")
        
        profile.interests.set(interests_qs)
        print(f"[Register] Successfully set interests on profile ID: {profile.id}")
        
        # --- Handle Social Accounts (Batch Save) ---
        social_accounts_raw = request.data.get('social_accounts')
        if social_accounts_raw:
            print(f"[Register] Processing social accounts...")
            try:
                import json
                if isinstance(social_accounts_raw, str):
                    try:
                        social_data = json.loads(social_accounts_raw)
                    except:
                        social_data = [] # Failed to parse
                elif isinstance(social_accounts_raw, list):
                    social_data = social_accounts_raw
                else:
                    social_data = []
                    
                for acc in social_data:
                    platform = acc.get('platform')
                    if not platform: continue
                    
                    # Normalize manual data
                    followers = acc.get('followers_count', 0)
                    try:
                        followers = int(float(str(followers).replace(',', '')))
                    except:
                        followers = 0
                    
                    username = acc.get('username', '')
                    profile_url = acc.get('profile_url', acc.get('link', ''))
                    
                    # Use update_or_create to prevent IntegrityError (and broken transactions)
                    SocialPlatformAccount.objects.update_or_create(
                        user=request.user,
                        platform=platform,
                        defaults={
                            'username': username or "User",
                            'profile_url': profile_url,
                            'followers_count': followers,
                            'is_verified': False
                        }
                    )
                print(f"[Register] Social accounts saved/updated.")
            except Exception as e:
                # If error happens here, we must Log it but NOT suppress it blindly if it broke the transaction?
                # Actually, update_or_create shouldn't break it. 
                # But if something else breaks, we should re-raise or ensure we don't return success if transaction is broken.
                print(f"[Register] Warning: Social accounts processing issue: {e}")
                # We can continue, as long as the DB transaction wasn't broken by a bad query.
        
        # --- Background OCR Task (Thread Pool) ---
        # Using a proper ThreadPoolExecutor allows us to QUEUE requests if many come at once.
        # This prevents the server from crashing due to too many threads.
        # The executor is defined globally at the top level of this file (see below) or we create a singleton.
        
        # We will define the task function outside the view to be clean (or keep it here for closure access).
        # Keeping it here for simplicity of 'p.id' access, but wrapped in a helper.
        
        # Call OCR task ONLY after transaction commits successfully
        transaction.on_commit(lambda: RegistrationSubmitView.submit_ocr_task(profile.id))
        
        # Update user status to PENDING
        request.user.submit_registration()
        
        return Response({
            'success': True,
            'message': 'Registration submitted successfully',
            'user': UserSerializer(request.user).data
        }, status=status.HTTP_201_CREATED)

    # Define Executor at class level (or module level)
    # max_workers=2 means only 2 OCRs run at once. Others wait in queue. High Performance & Safe.
    from concurrent.futures import ThreadPoolExecutor
    ocr_executor = ThreadPoolExecutor(max_workers=2, thread_name_prefix="OCR_Worker")

    @staticmethod
    def run_ocr_task(pid):
        """Static method for background OCR execution."""
        try:
            from django.db import close_old_connections
            import time
            close_old_connections()
            
            print(f"[OCR Queue] Starting Processing for Profile #{pid}")
            
            # Retry fetching profile (in case of replication lag or race condition)
            p = None
            for i in range(5):
                try:
                    p = InfluencerProfile.objects.get(pk=pid)
                    break
                except InfluencerProfile.DoesNotExist:
                    print(f"[OCR] Profile {pid} not found (attempt {i+1}). Waiting...")
                    time.sleep(1)
            
            if not p:
                print(f"[OCR] Error: Profile {pid} never appeared in DB.")
                return

            service = GeminiService()
            
            # 1. ID Card
            if p.id_card_front:
                try:
                    print(f"[OCR] Processing ID Card: {p.id_card_front.name}")
                    with p.id_card_front.open('rb') as f:
                        id_data = service.extract_document_data(f, 'id_card')
                        if id_data:
                            print(f"[OCR] ID Data: {id_data}")
                            if 'identification_number' in id_data:
                                p.identification_number = id_data['identification_number']
                except Exception as e:
                    print(f"[OCR] ID Card Failed: {e}")

            # 2. Bank Book
            if p.bank_book:
                try:
                    print(f"[OCR] Processing Bank Book: {p.bank_book.name}")
                    with p.bank_book.open('rb') as f:
                        bank_data = service.extract_document_data(f, 'bank_book')
                        if bank_data:
                            print(f"[OCR] Bank Data: {bank_data}")
                            if 'bank_name' in bank_data:
                                p.bank_name = bank_data['bank_name']
                            if 'account_number' in bank_data:
                                p.bank_account_number = bank_data['account_number']
                except Exception as e:
                    print(f"[OCR] Bank Book Failed: {e}")
            
            # Refresh connection before saving
            close_old_connections()
            
            # Simple retry mechanism for saving
            for i in range(3):
                try:
                    p.save()
                    print(f"[OCR] Successfully saved OCR data for Profile #{pid}")
                    break
                except Exception as save_err:
                    print(f"[OCR] Save failed (attempt {i+1}): {save_err}")
                    close_old_connections()
                    if i == 2: raise save_err
            
            print(f"[OCR Queue] Finished Task Profile #{pid}")
        except Exception as e:
            print(f"[OCR Queue] Task Failed: {e}")
            from django.db import close_old_connections
            close_old_connections()

    @classmethod
    def submit_ocr_task(cls, pid):
        cls.ocr_executor.submit(cls.run_ocr_task, pid)
        return True


from rest_framework.parsers import JSONParser

class ProfileUpdateView(APIView):
    """Update user profile personal info and work conditions."""
    permission_classes = [IsAuthenticated]
    parser_classes = (MultiPartParser, FormParser, JSONParser)
    
    @transaction.atomic
    def post(self, request):
        return self.put(request)

    @transaction.atomic
    def put(self, request):
        print(f"[ProfileUpdate] Content-Type: {request.content_type}")
        print(f"[ProfileUpdate] RAW DATA: {request.data}")
        
        if not hasattr(request.user, 'profile'):
            return Response(
                {'error': 'no_profile', 'message': 'Profile not found'},
                status=status.HTTP_404_NOT_FOUND
            )
        
        profile = request.user.profile
        from .serializers import ProfileUpdateSerializer
        
        serializer = ProfileUpdateSerializer(profile, data=request.data, partial=True)
        if not serializer.is_valid():
            print(f"[ProfileUpdate] Validation Errors: {serializer.errors}")
            return Response({
                'success': False,
                'message': 'ข้อมูลไม่ถูกต้อง',
                'errors': serializer.errors
            }, status=status.HTTP_400_BAD_REQUEST)

        # Validate mandatory documents (FRONTEND already ensures this, but BACKEND is the source of truth)
        has_id = bool(profile.id_card_front) or 'id_card_front' in request.FILES
        has_bank = bool(profile.bank_book) or 'bank_book' in request.FILES
        
        if not has_id or not has_bank:
            return Response({
                'success': False,
                'message': 'บัตรประชาชนและหน้าสมุดบัญชีเป็นข้อมูลที่จำเป็น',
                'details': {'id_card_front': not has_id, 'bank_book': not has_bank}
            }, status=status.HTTP_400_BAD_REQUEST)

        # Save profile
        profile = serializer.save()
        
        # SUPER DEFENSIVE FALLBACK: Explicitly set pricing fields if they are in data
        def force_price(val):
             if val is None or str(val).strip() == '': return None
             try:
                 # Remove commas and whitespace
                 clean_val = str(val).replace(',', '').strip()
                 return Decimal(clean_val)
             except Exception as e:
                 print(f"[ProfileUpdate] Price parse error for '{val}': {e}")
                 return None

        # Manual overrides to be 101% sure
        print(f"[ProfileUpdate] RECEIVING: boost_price={request.data.get('boost_price')}, original={request.data.get('original_file_price')}")
        
        if 'boost_price' in request.data:
            profile.boost_price = force_price(request.data.get('boost_price'))
            if profile.boost_price is not None:
                 profile.allow_boost = True
                 print(f"[ProfileUpdate] SET boost_price={profile.boost_price}, allow=True")

        if 'original_file_price' in request.data:
            profile.original_file_price = force_price(request.data.get('original_file_price'))
            if profile.original_file_price is not None:
                 profile.allow_original_file = True
                 print(f"[ProfileUpdate] SET original_file_price={profile.original_file_price}, allow=True")

        profile.save()
        profile.refresh_from_db()
        print(f"[ProfileUpdate] FINAL DB STATE: boost={profile.boost_price} ({type(profile.boost_price)}), allow={profile.allow_boost}")

        # Handle Interests (from serializer validated_data or request.data)
        interests_data = request.data.get('interests')
        if interests_data:
            import json
            try:
                if isinstance(interests_data, str):
                    interests_data = json.loads(interests_data)
                interests = Interest.objects.filter(id__in=interests_data)
                profile.interests.set(interests)
            except Exception as e:
                print(f"[ProfileUpdate] Interest error: {e}")

        # Handle Social Accounts
        social_data = request.data.get('social_accounts')
        if social_data:
            import json
            try:
                if isinstance(social_data, str):
                    social_data = json.loads(social_data)
                
                processed_platforms = []
                for acc in social_data:
                    platform = acc.get('platform')
                    if not platform: continue
                    
                    followers = acc.get('followers_count', 0)
                    try:
                        followers = int(float(str(followers).replace(',', '')))
                    except:
                        followers = 0
                        
                    SocialPlatformAccount.objects.update_or_create(
                        user=request.user,
                        platform=platform,
                        defaults={
                            'username': acc.get('username', 'User'),
                            'profile_url': acc.get('profile_url', acc.get('link', '')),
                            'followers_count': followers
                        }
                    )
                    processed_platforms.append(platform)
                
                # Sync: Delete accounts not in list
                SocialPlatformAccount.objects.filter(user=request.user).exclude(platform__in=processed_platforms).delete()
            except Exception as e:
                print(f"[ProfileUpdate] Social error: {e}")

        # RE-FETCH USER TO ENSURE LATEST DATA SERIALIZATION
        fresh_user = User.objects.get(id=request.user.id)
        return Response({
            'success': True,
            'message': 'Profile updated successfully',
            'debug_info': {
                'allow_boost': profile.allow_boost,
                'boost_price': str(profile.boost_price),
                'allow_original_file': profile.allow_original_file,
                'original_file_price': str(profile.original_file_price)
            },
            'user': UserWithProfileSerializer(fresh_user, context={'request': request}).data
        })


class ProfileDebugView(APIView):
    """Temporary endpoint to check DB state."""
    permission_classes = [IsAuthenticated]
    def get(self, request):
        p = request.user.profile
        return Response({
            'allow_boost': p.allow_boost,
            'boost_price': str(p.boost_price),
            'allow_original_file': p.allow_original_file,
            'original_file_price': str(p.original_file_price)
        })

class CampaignListView(ListAPIView):
    """
    List campaigns with sorting priority.
    
    Priority:
    1. Active (user is working on it)
    2. Pending Review (user submitted work)
    3. Open (accepting applications)
    4. Closed
    """
    
    permission_classes = [IsAuthenticated]
    serializer_class = CampaignListSerializer
    
    def get_queryset(self):
        # 🔄 Trigger automatic status updates based on deadlines
        try:
            Campaign.update_all_statuses()
        except Exception as e:
            print(f"⚠️ Failed to auto-update campaign statuses: {e}")

        # Check if user is approved
        if not self.request.user.is_approved:
            return Campaign.objects.none()
        
        # Get user's applications
        user_applications = CampaignApplication.objects.filter(
            user=self.request.user
        ).values('campaign_id', 'status')
        
        # Create a mapping of campaign_id -> status
        app_status_map = {app['campaign_id']: app['status'] for app in user_applications}
        
        # DEBUG: Print user applications
        print(f"[CampaignList] User: {self.request.user.id} ({self.request.user.email})")
        print(f"[CampaignList] Applications raw: {list(user_applications)}")
        print(f"[CampaignList] App Status Map: {app_status_map}")
        
        # Build queryset with sorting
        queryset = Campaign.objects.exclude(status='DRAFT')
        
        # Define Priority Groups
        # 1. Not Applied (Not in app_status_map AND Open) will be handled by default case + filter
        
        # 2. Working
        working_statuses = ['APPROVED', 'WORK_IN_PROGRESS', 'SCRIPT_APPROVED', 'DRAFT_APPROVED']
        
        # 3. Waiting
        waiting_statuses = ['WAITING', 'SUBMITTED_SCRIPT', 'SUBMITTED_DRAFT', 'SUBMITTED_FINAL']
        
        # 4. Completed (Waiting for Payment)
        completed_statuses = ['COMPLETED']

        # 5. Finished (Money Received)
        finished_statuses = ['PAYMENT_TRANSFERRED']
        
        # Get IDs for each priority group
        working_ids = [cid for cid, s in app_status_map.items() if s in working_statuses]
        waiting_ids = [cid for cid, s in app_status_map.items() if s in waiting_statuses]
        completed_ids = [cid for cid, s in app_status_map.items() if s in completed_statuses]
        finished_ids = [cid for cid, s in app_status_map.items() if s in finished_statuses]
        
        print(f"[CampaignList] Waiting IDs: {waiting_ids}")
        print(f"[CampaignList] Working IDs: {working_ids}")
        
        # Any other ID in map (e.g. REJECTED) or not in map is handled below
        
        queryset = queryset.annotate(
            priority=Case(
                # Priority 1: Working / Active Revision
                When(id__in=working_ids, then=Value(1)),
                
                # Priority 2: Waiting (Submission sent)
                When(id__in=waiting_ids, then=Value(2)),
                
                # Priority 3: Completed (Wait for Payment)
                When(id__in=completed_ids, then=Value(3)),
                
                # Priority 4: Open for Applications (Not applied)
                When(~Q(id__in=app_status_map.keys()) & Q(status='OPEN'), then=Value(4)),

                # Priority 5: Finished / Rejected / Closed
                default=Value(5),
                output_field=IntegerField()
            )
        ).order_by('priority', 'application_deadline')
        
        # Filter by status if requested (Frontend might want to filter specific tabs)
        # Check both query_params (GET) and data (POST)
        status_filter = self.request.data.get('status') or self.request.query_params.get('status')
        print(f"[CampaignList] Status Filter: {status_filter}")
        
        if status_filter == 'active':
             # Show Working + Waiting + Completed (Wait for Payment)
             ids_to_include = working_ids + waiting_ids + completed_ids
             print(f"[CampaignList] Filtering active. IDs: {ids_to_include}")
             queryset = queryset.filter(id__in=ids_to_include)
        elif status_filter == 'history':
             # Show Finished (Payment Transferred) + Rejected
             # COMPLETED (Wait for Payment) stays in 'active' tab until transferred
             history_ids = finished_ids + [cid for cid, s in app_status_map.items() if s == 'REJECTED']
             print(f"[CampaignList] Filtering history. IDs: {history_ids}")
             queryset = queryset.filter(id__in=history_ids)
        
        return queryset
    
    def get_serializer_context(self):
        context = super().get_serializer_context()
        context['request'] = self.request
        return context
    
    def post(self, request, *args, **kwargs):
        """Support POST for filtering campaigns (hides params from URL)."""
        return self.list(request, *args, **kwargs)


class CampaignDetailView(RetrieveAPIView):
    """Get campaign detail with timeline."""
    
    permission_classes = [IsAuthenticated]
    serializer_class = CampaignDetailSerializer
    queryset = Campaign.objects.all()
    lookup_field = 'uuid' # Changed from 'pk' or default to 'uuid'

    def get_object(self):
        """
        Returns the object the view is displaying.
        You may want to override this if you need to provide non-standard
        queryset lookups.
        """
        queryset = self.filter_queryset(self.get_queryset())

        # Perform the lookup using kwargs
        lookup_url_kwarg = self.lookup_url_kwarg or self.lookup_field

        filter_kwargs = {self.lookup_field: self.kwargs[lookup_url_kwarg]}
        
        # Fallback to ID if not valid UUID
        try:
            import uuid
            uuid.UUID(str(self.kwargs[lookup_url_kwarg]))
        except ValueError:
             # If not UUID, assume it's an ID (legacy support for admin/internal)
             filter_kwargs = {'id': self.kwargs[lookup_url_kwarg]}

        obj = get_object_or_404(queryset, **filter_kwargs)

        # May raise a permission denied
        self.check_object_permissions(self.request, obj)

        return obj
    permission_classes = [IsAuthenticated]
    serializer_class = CampaignDetailSerializer
    queryset = Campaign.objects.exclude(status='DRAFT')
    
    def get(self, request, *args, **kwargs):
        # Story 3.1: Block Unapproved Users from Jobs
        if not request.user.is_approved:
            from rest_framework.exceptions import PermissionDenied
            raise PermissionDenied("You must be approved before viewing campaign details.")
        return super().get(request, *args, **kwargs)

    def get_serializer_context(self):
        context = super().get_serializer_context()
        context['request'] = self.request
        return context


class CampaignApplyView(APIView):
    """Apply to a campaign."""
    
    permission_classes = [IsAuthenticated]
    
    def post(self, request, uuid):
        # Check if user is approved
        if not request.user.is_approved:
            return Response(
                {'error': 'user_not_approved', 'message': 'You must be approved before applying to campaigns'},
                status=status.HTTP_403_FORBIDDEN
            )
        
        # Get campaign
        try:
            # Try UUID first
            import uuid as uuid_lib
            try:
                uuid_obj = uuid_lib.UUID(str(uuid))
                campaign = Campaign.objects.get(uuid=uuid)
            except (ValueError, Campaign.DoesNotExist):
                # Fallback to ID
                campaign = Campaign.objects.get(id=uuid)
        except Campaign.DoesNotExist:
            return Response(
                {'error': 'not_found', 'message': 'Campaign not found'},
                status=status.HTTP_404_NOT_FOUND
            )
        
        # Check if campaign is open
        if not campaign.is_open:
            return Response(
                {'error': 'campaign_closed', 'message': 'This campaign is not accepting applications'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Check if already applied
        if CampaignApplication.objects.filter(campaign=campaign, user=request.user).exists():
            return Response(
                {'error': 'already_applied', 'message': 'You have already applied to this campaign'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Create application
        note = request.data.get('note', '')
        application = CampaignApplication.objects.create(
            campaign=campaign,
            user=request.user,
            application_note=note
        )
        
        return Response({
            'success': True,
            'application': {
                'id': application.id,
                'campaign_id': campaign.id,
                'status': application.status,
                'applied_at': application.applied_at
            }
        }, status=status.HTTP_201_CREATED)


class UserApplicationsView(ListAPIView):
    """List user's campaign applications."""
    
    permission_classes = [IsAuthenticated]
    serializer_class = ApplicationSerializer
    
    def get_queryset(self):
        return CampaignApplication.objects.filter(
            user=self.request.user
        ).select_related('campaign')


class SubmitWorkView(APIView):
    """Submit work for a campaign application stage."""
    
    permission_classes = [IsAuthenticated]
    
    def post(self, request, pk):
        # Get application
        try:
            application = CampaignApplication.objects.get(
                pk=pk,
                user=request.user
            )
        except CampaignApplication.DoesNotExist:
            return Response(
                {'error': 'not_found', 'message': 'Application not found'},
                status=status.HTTP_404_NOT_FOUND
            )
        
        serializer = SubmitWorkSerializer(data=request.data, context={'request': request})
        if not serializer.is_valid():
            return Response(
                {'success': False, 'errors': serializer.errors},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        data = serializer.validated_data
        submission_link = data.get('link')
        
        # Validate that user can submit for this stage
        allowed_stages = {
            'WORK_IN_PROGRESS': 'script',
            'APPROVED': 'script',
            'REVISE_SCRIPT': 'script',
            'SCRIPT_APPROVED': 'draft',
            'REVISE_DRAFT': 'draft',
            'DRAFT_APPROVED': 'final',
            'REVISE_FINAL': 'final',
            'FINAL_APPROVED': 'insight',
            'REVISE_INSIGHT': 'insight',
        }
        
        expected_stage = allowed_stages.get(application.status)
        if data['stage'] != expected_stage:
            return Response(
                {
                    'error': 'invalid_stage',
                    'message': f'You can only submit {expected_stage} at this time'
                },
                status=status.HTTP_400_BAD_REQUEST
            )
        
        if data['stage'] == 'insight':
            insight_urls = []
            
            # 1. Handle Retention of Existing Files
            # Retrieve list of files user wants to keep
            kept_urls = request.data.getlist('kept_files') if 'kept_files' in request.data else []
            # Ensure kept_urls are actually from our own storage domain (basic security check optional)
            
            # Identify files to delete (Old files NOT in kept_urls)
            old_files = application.insight_files or []
            if application.insight_image and application.insight_image not in old_files:
                old_files.append(application.insight_image)
            
            for old_url in old_files:
                if old_url not in kept_urls:
                    # DELETE this file
                    try:
                        # Extract relative path from URL (assuming default_storage)
                        # This works for standard Django storage. Adjust if using S3/GCS with signed URLs logic.
                        # Simple approach: If it stores relative path in DB, great. If full URL, need to parse.
                        # Here we assume application.insight_files stores the Relative Path or Full URL depending on setup.
                        # Based on line 1048 `default_storage.url(saved_path)`, it stores the URL.
                        # To delete, we need the path.
                        
                        # Heuristic: Try to find the path suffix
                        # Unquote URL first
                        from urllib.parse import unquote
                        decoded_url = unquote(old_url)
                        
                        if 'campaigns/' in decoded_url:
                            # Extract part after media URL or just search for campaigns/...
                            rel_path = decoded_url[decoded_url.find('campaigns/'):]
                            if default_storage.exists(rel_path):
                                print(f"[Insight Cleanup] Deleting {rel_path}")
                                default_storage.delete(rel_path)
                    except Exception as del_err:
                        print(f"[Insight Cleanup Error] Failed to delete {old_url}: {del_err}")

            # Add kept files to the new list
            insight_urls.extend(kept_urls)

            # 2. Handle New Uploads
            if 'file' in request.FILES:
                files = [request.FILES['file']]
            elif 'files' in request.FILES:
                files = request.FILES.getlist('files')
            else:
                files = []

            if files:
                try:
                    from PIL import Image
                    from io import BytesIO
                    from django.core.files.uploadedfile import InMemoryUploadedFile
                    from django.core.files.storage import default_storage
                    import time
                    import uuid
                    import sys

                    print(f"[Insight Upload] Processing {len(files)} new files...")

                    for idx, image_file in enumerate(files):
                        # Open image
                        im = Image.open(image_file)
                        
                        # Convert to RGB if RGBA
                        if im.mode == 'RGBA':
                            im = im.convert('RGB')
                        
                        # Resize if too large (max 1920px width)
                        if im.width > 1920:
                            ratio = 1920 / im.width
                            new_height = int(im.height * ratio)
                            im = im.resize((1920, new_height), Image.Resampling.LANCZOS)
                        
                        # Compress
                        output = BytesIO()
                        im.save(output, format='JPEG', quality=70, optimize=True)
                        output.seek(0)
                        
                        # Create new Django file
                        file_size = output.getbuffer().nbytes
                        file_ext = image_file.name.split('.')[-1]
                        
                        # UNIQUE NAMING: user_{id}_insight_{timestamp}_{uuid}.jpg
                        filename = f"user_{application.user.id}_insight_{int(time.time())}_{uuid.uuid4().hex[:8]}.jpg"
                        
                        new_image = InMemoryUploadedFile(
                            output,
                            'ImageField',
                            filename,
                            'image/jpeg',
                            file_size,
                            None
                        )
                        
                        # Save to storage
                        file_path = f"campaigns/{application.campaign.id}/insights/{filename}"
                        
                        saved_path = default_storage.save(file_path, new_image)
                        
                        # Get URL
                        url = default_storage.url(saved_path)
                        
                        # Only append cache-busting if we are sure it's not a signed URL
                        # For GCS/S3, better to rely on their caching policies or just use the URL as is.
                        # We will remove the manual ?t= appending to be safe.
                        
                        insight_urls.append(url)

                except Exception as e:
                    print(f"[Insight File Processing Error] {e}")
                    import traceback
                    traceback.print_exc()
                    # Continue even if one file fails? Or fail whole request?
                    # For now, let's just log and continue, but strict strictness depends on reqs.
                    # Ideally we want to fail if upload fails.
                    pass


            # Update Application Data
            # Store primary image in legacy field for backward compatibility
            if insight_urls:
                submission_link = insight_urls[0] 
                application.insight_image = insight_urls[0]
                application.insight_files = insight_urls # New JSON Field
                
                # IMPORTANT: Update stage status
                application.submit_work(data['stage'], submission_link, notes=data.get('notes', ''))
                
                # Explicitly save changes before AI analysis
                application.save(update_fields=['insight_image', 'insight_files'])
            else:
                 # Fallback if somehow empty, though validation should prevent this
                 submission_link = ""
                 # Still proceed to submit work to trigger status update
                 application.submit_work(data['stage'], submission_link, notes=data.get('notes', ''))
                    
            # --- REAL AI ANALYSIS (Gemini Vision Multi-File) ---
            try:
                from .services.gemini import GeminiService
                
                print(f"[AI Analysis] Triggering Multi-File Analysis for {len(insight_urls)} images")
                ai_result = GeminiService.analyze_insight_images(insight_urls)
                        
                if ai_result:
                    application.insight_data = ai_result
                    application.save(update_fields=['insight_data', 'insight_files', 'insight_image'])
                    print(f"[AI Analysis] Saved REAL data for {application.id}")

                    # --- NEW: Save Dedicated Metrics for Reporting ---
                    try:
                        from apps.campaigns.models import CampaignInsightMetric
                        
                        # 1. Extract Raw Metrics (Handle K/M suffixes via helper or assume raw numbers if cleaned)
                        def clean_num(val):
                            if val is None: return 0
                            if isinstance(val, (int, float)): return int(val)
                            s = str(val).lower().replace(',', '').strip()
                            if not s: return 0
                            try:
                                if 'k' in s: return int(float(s.replace('k', '')) * 1000)
                                if 'm' in s: return int(float(s.replace('m', '')) * 1000000)
                                return int(float(s))
                            except:
                                return 0

                        metrics = ai_result.get('metrics', {})
                        views = clean_num(metrics.get('views', 0))
                        likes = clean_num(metrics.get('likes', 0))
                        comments = clean_num(metrics.get('comments', 0))
                        shares = clean_num(metrics.get('shares', 0))
                        
                        # 2. Calculate KPI
                        eng_rate = 0.0
                        if views > 0:
                            eng_rate = ((likes + comments + shares) / views) * 100
                        
                        cpv = 0.0
                        budget = 0.0
                        try:
                            budget = float(application.campaign.budget)
                        except: pass

                        if views > 0 and budget > 0:
                            cpv = budget / views
                            
                        # 3. Save to DB
                        CampaignInsightMetric.objects.update_or_create(
                            application=application,
                            defaults={
                                'total_views': views,
                                'total_likes': likes,
                                'total_comments': comments,
                                'total_shares': shares,
                                'engagement_rate': round(eng_rate, 2),
                                'cost_per_view': round(cpv, 4)
                            }
                        )
                        print(f"[Metrics] Saved Metrics for App #{application.id}: Views={views}, ER={eng_rate}%, CPV={cpv}")
                        
                    except Exception as m_err:
                        # FAIL OPEN: Do not crash the submission if metrics fail
                        print(f"[Metrics] Failed to save DB metrics (Ignored): {m_err}")
                        import traceback
                        traceback.print_exc()
                else:
                    print(f"[AI Analysis] Failed to get result from Gemini")
                    # Still save file paths even if AI fails
                    application.save(update_fields=['insight_files', 'insight_image'])
                    
            except Exception as ai_err:
                print(f"[AI Analysis] Critical Error: {ai_err}")
                import traceback
                traceback.print_exc()
                # Ensure files are saved at minimum
                application.save(update_fields=['insight_files', 'insight_image'])
            
            return Response({'success': True, 'message': 'Insight submitted successfully', 'link': submission_link, 'files_count': len(insight_urls)})





        # Submit work
        application.submit_work(
            stage=data['stage'],
            link=submission_link,
            notes=data.get('notes', '')
        )
        
        # Get latest submission data for response
        submission_list = application.submission_data.get(data['stage'], [])
        latest_submission = submission_list[-1] if submission_list else {}

        return Response({
            'success': True,
            'submission': {
                'stage': data['stage'],
                'link': submission_link,
                'submitted_at': latest_submission.get('submitted_at'),
                'status': 'pending'
            },
            'application_status': application.status
        })


class ValidateDriveLinkView(APIView):
    """Validate Google Drive link accessibility."""
    
    permission_classes = [IsAuthenticated]
    
    def post(self, request):
        serializer = ValidateDriveLinkSerializer(data=request.data)
        if not serializer.is_valid():
            return Response(
                {'success': False, 'errors': serializer.errors},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        link = serializer.validated_data['link']
        result = GoogleDriveValidator.validate(link)
        
        return Response(result)


class SendMessageView(APIView):
    """
    Send LINE messages via Messaging API.
    
    This is a centralized endpoint for sending LINE messages.
    Frontend can call this API to send text or flex messages.
    """
    
    permission_classes = [IsAuthenticated]
    
    def post(self, request):
        from apps.users.services import LineMessagingService
        
        message_type = request.data.get('type', 'text')
        line_user_id = request.data.get('line_user_id')
        
        # If no line_user_id provided, use current user's
        if not line_user_id:
            line_user_id = request.user.line_user_id
        
        if not line_user_id:
            return Response(
                {'success': False, 'error': 'line_user_id is required'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        if message_type == 'text':
            text = request.data.get('text')
            if not text:
                return Response(
                    {'success': False, 'error': 'text is required for text message'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            result = LineMessagingService.send_text_message(line_user_id, text)
            
        elif message_type == 'flex':
            alt_text = request.data.get('altText', 'Message from Influencer Platform')
            contents = request.data.get('contents')
            if not contents:
                return Response(
                    {'success': False, 'error': 'contents is required for flex message'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            result = LineMessagingService.send_flex_message(line_user_id, alt_text, contents)
            
        elif message_type == 'raw':
            # Raw messages - user provides complete message array
            messages = request.data.get('messages')
            if not messages:
                return Response(
                    {'success': False, 'error': 'messages is required for raw message'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            result = LineMessagingService.send_push_message(line_user_id, messages)
            
        else:
            return Response(
                {'success': False, 'error': f'Unknown message type: {message_type}'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        if result.get('success'):
            return Response({'success': True, 'message': 'Message sent successfully'})
        else:
            return Response(
                {'success': False, 'error': result.get('error', 'Failed to send message')},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )


class SocialAccountsListView(APIView):
    """List user's connected social media accounts."""
    
    permission_classes = [IsAuthenticated]
    
    def get(self, request):
        accounts = SocialPlatformAccount.objects.filter(user=request.user)
        serializer = SocialPlatformAccountSerializer(accounts, many=True)
        return Response({
            'success': True,
            'accounts': serializer.data
        })


class SocialConnectView(APIView):
    """Connect a new social media account by fetching profile data."""
    
    permission_classes = [IsAuthenticated]
    
    @transaction.atomic
    @transaction.atomic
    def post(self, request):
        serializer = SocialConnectSerializer(data=request.data)
        if not serializer.is_valid():
            return Response(
                {'success': False, 'errors': serializer.errors},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        platform = serializer.validated_data['platform']
        profile_url = serializer.validated_data['profile_url']
        
        # Check if account already connected
        existing = SocialPlatformAccount.objects.filter(
            user=request.user,
            platform=platform
        ).first()
        
        if existing:
            return Response(
                {'success': False, 'error': f'{platform.title()} account already connected'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        profile_data = {}
        
        # 1. Manual Input Override (if frontend sends followers_count)
        if request.data.get('followers_count'):
            username = request.data.get('username', '')
            
            # Try to extract username from URL if not provided or looks like URL
            if not username or 'http' in username or '/' in username:
                try:
                    # Quick logic to extract username based on platform
                    if platform == 'tiktok':
                        import re
                        match = re.search(r'tiktok\.com/(@?[\w.-]+)', profile_url)
                        if match: username = f"@{match.group(1).replace('@', '')}"
                    elif platform == 'instagram':
                        import re
                        match = re.search(r'instagram\.com/([a-zA-Z0-9_\.]+)', profile_url)
                        if match: username = f"@{match.group(1)}"
                    elif platform == 'facebook':
                        import re
                        match = re.search(r'facebook\.com/([a-zA-Z0-9.]+)', profile_url)
                        if match: username = match.group(1)
                    elif platform == 'youtube':
                        # YouTube is complex, just use what we have or placeholder
                        username = "YouTube Channel"
                except:
                    pass
            
            profile_data = {
                'success': True,
                'data': {
                    'platform': platform,
                    'username': username or "User",
                    'followers_count': request.data.get('followers_count', 0),
                    'profile_picture_url': '',
                    'is_verified': False
                }
            }
            
        else:
            # 2. Automatic Fetch
            print(f"[SocialConnect] Fetching profile for {platform} at {profile_url}")
            profile_data = SocialPlatformService.fetch_profile(platform, profile_url)
            print(f"[SocialConnect] Fetch result success: {profile_data.get('success')}")
        
        # 3. Fallback: If fetch failed, try to just save the link (Regex Extraction)
        if not profile_data.get('success'):
            print(f"[SocialConnect] Fetch Failed: {profile_data.get('error')}. Attempting Fallback...")
            
            # Attempt to extract minimally required data (username) from URL
            fallback_username = None
            try:
                if platform == 'tiktok':
                    from apps.influencers.services import TikTokLocalService
                    fallback_username = TikTokLocalService.extract_username(profile_url)
                    if fallback_username: fallback_username = f"@{fallback_username}"
                elif platform == 'instagram':
                    from apps.influencers.services import InstaloaderService
                    fallback_username = InstaloaderService.extract_username(profile_url)
                    if fallback_username: fallback_username = f"@{fallback_username}"
                elif platform == 'facebook':
                    from apps.influencers.services import FacebookLocalService
                    fallback_username = FacebookLocalService.extract_username(profile_url)
                elif platform == 'youtube':
                    from apps.influencers.services import YouTubeService
                    type_, val = YouTubeService.extract_channel_id_or_handle(profile_url)
                    fallback_username = val
            except Exception as e:
                print(f"[SocialConnect] Fallback Regex Failed: {e}")
            
            if fallback_username:
                print(f"[SocialConnect] Fallback successful. Username: {fallback_username}")
                profile_data = {
                    'success': True,
                    'data': {
                        'platform': platform,
                        'username': fallback_username,
                        'profile_url': profile_url,
                        'profile_picture_url': '',
                        'followers_count': 0, # Cannot verify
                        'following_count': 0,
                        'posts_count': 0,
                        'is_verified': False
                    }
                }
            else:
                # Truly failed
                return Response(
                    {'success': False, 'error': profile_data.get('error', 'Failed to fetch profile data and could not verify URL')},
                    status=status.HTTP_400_BAD_REQUEST
                )
        
        final_data = profile_data['data']
        
        # Create the social account record
        account = SocialPlatformAccount.objects.create(
            user=request.user,
            platform=final_data['platform'],
            platform_user_id=final_data.get('platform_user_id', ''),
            username=final_data.get('username', ''),
            profile_url=profile_url,
            profile_picture_url=final_data.get('profile_picture_url', ''),
            followers_count=final_data.get('followers_count', 0),
            following_count=final_data.get('following_count', 0),
            posts_count=final_data.get('posts_count', 0),
            is_verified=final_data.get('is_verified', False)
        )
        
        return Response({
            'success': True,
            'message': f'{platform.title()} account connected successfully',
            'account': SocialPlatformAccountSerializer(account).data
        }, status=status.HTTP_201_CREATED)


class SocialDisconnectView(APIView):
    """Disconnect a social media account."""
    
    permission_classes = [IsAuthenticated]
    
    def delete(self, request, pk):
        try:
            account = SocialPlatformAccount.objects.get(
                pk=pk,
                user=request.user
            )
        except SocialPlatformAccount.DoesNotExist:
            return Response(
                {'success': False, 'error': 'Account not found'},
                status=status.HTTP_404_NOT_FOUND
            )
        
        platform = account.get_platform_display()
        account.delete()
        
        return Response({
            'success': True,
            'message': f'{platform} account disconnected'
        })


class SocialSyncView(APIView):
    """Refresh/sync data for a connected social account."""
    
    permission_classes = [IsAuthenticated]
    
    def post(self, request, pk):
        try:
            account = SocialPlatformAccount.objects.get(
                pk=pk,
                user=request.user
            )
        except SocialPlatformAccount.DoesNotExist:
            return Response(
                {'success': False, 'error': 'Account not found'},
                status=status.HTTP_404_NOT_FOUND
            )
        
        # Fetch fresh data
        result = SocialPlatformService.fetch_profile(account.platform, account.profile_url)
        
        if not result.get('success'):
            return Response(
                {'success': False, 'error': result.get('error', 'Failed to sync profile data')},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        profile_data = result['data']
        
        # Update account data
        account.platform_user_id = profile_data.get('platform_user_id', account.platform_user_id)
        account.username = profile_data.get('username', account.username)
        account.profile_picture_url = profile_data.get('profile_picture_url', account.profile_picture_url)
        account.followers_count = profile_data.get('followers_count', account.followers_count)
        account.following_count = profile_data.get('following_count', account.following_count)
        account.posts_count = profile_data.get('posts_count', account.posts_count)
        account.is_verified = profile_data.get('is_verified', account.is_verified)
        account.save()
        
        return Response({
            'success': True,
            'message': 'Account data synced successfully',
            'account': SocialPlatformAccountSerializer(account).data
        })


class ImageProxyView(APIView):
    """
    Proxy an image URL to bypass hotlinking protection or Referer checks.
    Does not store anything on the server.
    """
    permission_classes = [permissions.AllowAny] # Allow anyone to view profile images
    
    def get(self, request):
        url = request.query_params.get('url')
        if not url:
            return Response({'error': 'URL parameter is required'}, status=status.HTTP_400_BAD_REQUEST)
        
        # Check if URL is internal (relative or same host)
        # If so, redirect to it directly to avoid loopback issues and let Nginx/WhiteNoise handle it.
        # This fixes "Image broken but view full size works" issues.
        is_local = False
        if url.startswith('/'):
            is_local = True
        elif request.get_host() in url:
            is_local = True
            
        if is_local:
            return HttpResponseRedirect(url)

        try:
            # Mask the Referer to bypass protection
            resp = requests.get(url, stream=True, timeout=10, headers={
                'User-Agent': 'Mozilla/5.0'
            })
            
            if resp.status_code != 200:
                logger.error(f"[ImageProxy] Failed to fetch {url}: {resp.status_code}")
                return Response({'error': f'Failed to fetch image: {resp.status_code}'}, status=resp.status_code)
                
            # Stream the response back to the client
            content_type = resp.headers.get('Content-Type', 'image/jpeg')
            return StreamingHttpResponse(resp.iter_content(chunk_size=8192), content_type=content_type)
            
        except Exception as e:
             logger.error(f"[ImageProxy] Error fetching {url}: {e}")
             return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class InfluencerApprovalListView(ListAPIView):
    """
    Consolidated list view for Admin Influencer Approval Dashboard.
    Returns users with PENDING/REJECTED status by default.
    """
    permission_classes = [permissions.IsAdminUser]
    serializer_class = InfluencerApprovalSerializer

    def get_queryset(self):
        status_filter = self.request.query_params.get('status', 'PENDING')
        if status_filter == 'ALL':
            return User.objects.exclude(status='NEW').order_by('-created_at')
        return User.objects.filter(status=status_filter).order_by('-created_at')


class BulkInfluencerActionView(APIView):
    """
    Bulk approve or reject influencers.
    Expects: { "user_ids": [1, 2, 3], "action": "APPROVE" | "REJECT", "reason": "optional" }
    """
    permission_classes = [permissions.IsAdminUser]

    @transaction.atomic
    def post(self, request):
        user_ids = request.data.get('user_ids', [])
        action = request.data.get('action')
        reason = request.data.get('reason', '')

        if not user_ids or not action:
            return Response(
                {'error': 'invalid_data', 'message': 'user_ids and action are required'},
                status=status.HTTP_400_BAD_REQUEST
            )

        users = User.objects.filter(id__in=user_ids)
        if not users.exists():
            return Response(
                {'error': 'not_found', 'message': 'No users found with provided IDs'},
                status=status.HTTP_404_NOT_FOUND
            )

        updated_count = 0
        for user in users:
            if action == 'APPROVE':
                user.approve()
                log_action(request.user, 'approve_user', user)
                updated_count += 1
            elif action == 'REJECT':
                user.reject(reason=reason)
                log_action(request.user, 'reject_user', user, {'reason': reason})
                updated_count += 1

        return Response({
            'success': True,
            'message': f'Successfully updated {updated_count} influencers',
            'updated_count': updated_count
        })








class AdminCampaignListView(ListAPIView):
    """
    List all campaigns for admin with summary stats.
    """
    permission_classes = [permissions.IsAdminUser]
    serializer_class = CampaignListSerializer # Reuse or create specific one
    
    def get_queryset(self):
        return Campaign.objects.all().order_by('-created_at')


class AdminCampaignDetailView(APIView):
    """
    Get campaign detail with all participants for admin management.
    """
    permission_classes = [permissions.IsAdminUser]
    
    def get(self, request, uuid):
        try:
            # Try UUID first
            import uuid as uuid_lib
            try:
                uuid_obj = uuid_lib.UUID(str(uuid))
                campaign = Campaign.objects.get(uuid=uuid)
            except (ValueError, Campaign.DoesNotExist):
                # Fallback to ID
                campaign = Campaign.objects.get(id=uuid)
        except Campaign.DoesNotExist:
            return Response({'error': 'not_found'}, status=404)
            
        # Get all applications
        # Added insight_metric select_related to query
        applications = CampaignApplication.objects.filter(campaign=campaign).select_related('user', 'user__profile', 'insight_metric').prefetch_related('user__profile__interests')
        
        # Group by status/stage
        participants = []
        for app in applications:
            profile = app.user.profile if hasattr(app.user, 'profile') else None
            participants.append({
                'id': app.id,
                'user_id': app.user.id,
                'display_name': app.user.display_name,
                'picture_url': app.user.picture_url,
                'status': app.status,
                'current_stage': app.current_stage,
                'applied_at': app.applied_at,
                'updated_at': app.updated_at,
                'submissions': app.submission_data,
                # Insight data (simplified)
                'insight_files': app.insight_files,
                'insight_image': app.insight_image,
                'insight_data': app.insight_data,
                'insight_submitted_at': app.insight_submitted_at.isoformat() if app.insight_submitted_at else None,
                'insight_feedback': app.insight_feedback,
                'insight_note': app.insight_note,
                'payment_slip': app.payment_slip.url if app.payment_slip else None,
                # New Metrics from dedicated table
                'metrics': {
                    'total_views': getattr(app.insight_metric, 'total_views', 0) if hasattr(app, 'insight_metric') else 0,
                    'total_likes': getattr(app.insight_metric, 'total_likes', 0) if hasattr(app, 'insight_metric') else 0,
                    'total_comments': getattr(app.insight_metric, 'total_comments', 0) if hasattr(app, 'insight_metric') else 0,
                    'total_shares': getattr(app.insight_metric, 'total_shares', 0) if hasattr(app, 'insight_metric') else 0,
                    'engagement_rate': getattr(app.insight_metric, 'engagement_rate', 0.0) if hasattr(app, 'insight_metric') else 0.0,
                    'cost_per_view': getattr(app.insight_metric, 'cost_per_view', 0.0) if hasattr(app, 'insight_metric') else 0.0,
                },
                'profile': {
                     'phone': profile.phone if profile else '',
                     'fullname': profile.full_name_th if profile else '',
                     'email': profile.email if profile else '',
                     'interests': [
                         {'id': i.id, 'name': i.name_th, 'icon': i.icon} 
                         for i in profile.interests.all()
                     ] if profile else [],
                     'line_id': app.user.line_user_id
                },
                'social_accounts': [
                    {
                        'platform': acc.platform,
                        'username': acc.username,
                        'profile_url': acc.profile_url,
                        'followers_count': acc.followers_count,
                        'followers_formatted': acc.followers_formatted
                    } for acc in app.user.social_accounts.all()
                ]
            })
            
        # Use serializer for full details to ensure consistency
        serializer = CampaignUpdateSerializer(campaign, context={'request': request})
        campaign_full = serializer.data

        return Response({
            'campaign': {
                'id': campaign.id,
                'title': campaign.title,
                'brand_name': campaign.brand_name,
                'status': campaign.status,
                'budget': campaign.budget,
                'show_slip_to_client': campaign.show_slip_to_client,
            },
            'campaign_full': campaign_full,
            'participants': participants
        })

    def patch(self, request, uuid):
        try:
            # Try UUID first
            import uuid as uuid_lib
            try:
                uuid_obj = uuid_lib.UUID(str(uuid))
                campaign = Campaign.objects.get(uuid=uuid)
            except (ValueError, Campaign.DoesNotExist):
                # Fallback to ID
                campaign = Campaign.objects.get(id=uuid)
        except Campaign.DoesNotExist:
            return Response({'error': 'not_found'}, status=404)
        
        serializer = CampaignUpdateSerializer(campaign, data=request.data, partial=True)
        if serializer.is_valid():
            campaign = serializer.save()
            
            # Sanitize log data: convert QueryDict to dict and remove Files
            log_details = {}
            if hasattr(request.data, 'dict'):
                log_details = request.data.dict()
            else:
                log_details = dict(request.data)
                
            # Filter out file objects to keep logs JSON serializable
            log_details = {k: v for k, v in log_details.items() if not hasattr(v, 'read')}

            log_action(
                actor=request.user,
                action_type='edit_campaign',
                target=campaign,
                details={'changes': log_details},
                ip_address=request.META.get('REMOTE_ADDR')
            )
            return Response(serializer.data)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    def delete(self, request, uuid):
        """Delete a campaign with activity logging."""
        try:
            # Try UUID first
            import uuid as uuid_lib
            try:
                uuid_obj = uuid_lib.UUID(str(uuid))
                campaign = Campaign.objects.get(uuid=uuid)
            except (ValueError, Campaign.DoesNotExist):
                # Fallback to ID
                campaign = Campaign.objects.get(id=uuid)
        except Campaign.DoesNotExist:
            return Response({'error': 'not_found'}, status=404)
        
        # Store campaign info for logging before deletion
        campaign_title = campaign.title
        campaign_id = campaign.id
        campaign_brand = campaign.brand_name
        
        # Log the deletion action BEFORE deleting
        log_action(
            actor=request.user,
            action_type='delete_campaign',
            target=None,  # Campaign will be deleted, so we pass None
            details={
                'campaign_id': campaign_id,
                'campaign_title': campaign_title,
                'brand_name': campaign_brand,
                'deleted_by': request.user.display_name or request.user.username
            },
            ip_address=request.META.get('REMOTE_ADDR')
        )
        
        # Delete the campaign
        campaign.delete()
        
        return Response({
            'success': True,
            'message': f'Campaign "{campaign_title}" has been deleted successfully'
        })


class AdminWorkReviewView(APIView):
    """
    Review a specific stage of a campaign application.
    """
    permission_classes = [permissions.IsAdminUser]

    def post(self, request, pk):
        action = request.data.get('action') # APPROVE or REJECT (Revision)
        stage = request.data.get('stage')
        feedback = request.data.get('feedback', '')

        if not action or not stage:
            return Response({'error': 'Action and stage are required'}, status=400)

        try:
            app = CampaignApplication.objects.get(pk=pk)
            
            if action == 'APPROVE':
                app.approve_stage(stage, feedback)
                log_action(request.user, 'approve_work', app, {'stage': stage, 'feedback': feedback})
            elif action == 'REJECT':
                app.request_revision(stage, feedback)
                log_action(request.user, 'reject_work', app, {'stage': stage, 'feedback': feedback})
            else:
                return Response({'error': 'Invalid action'}, status=400)
                
            return Response({
                'success': True, 
                'status': app.status,
                'submission': app.submission_data.get(stage, {})
            })
        except CampaignApplication.DoesNotExist:
            return Response({'error': 'Not found'}, status=404)


class AdminExportParticipantsView(APIView):
    """
    Export campaign participants to CSV (Excel compatible).
    Matches user requested format: No, TikTok Name, Link, Follower, etc.
    """
    permission_classes = [permissions.IsAdminUser]
    
    def clean_text(self, text):
        if not text:
            return ""
        import re
        # Remove sequences of 2 or more question marks
        text = re.sub(r'\?{2,}', '', text)
        # Remove single question marks at start or end
        text = text.strip('?')
        return text.strip()
    
    def get(self, request, uuid):
        import csv
        from django.http import HttpResponse
        from apps.influencers.models import SocialPlatformAccount
        
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="campaign_{uuid}_participants.csv"'
        response.write(u'\ufeff'.encode('utf8')) # BOM for Excel
        
        writer = csv.writer(response)
        # Columns matching user request + Application ID for system
        headers = [
            'Application ID', 'No', 
            'TikTok Name', 'Link TikTok', 'Follower', 
            'Brand Approved', 
            'Rate Card', 'Condition', 
            'Action (APPROVE/REJECT)', 'Note'
        ]
        writer.writerow(headers)
        
        import uuid as uuid_lib
        try:
            # Try UUID first
            uuid_obj = uuid_lib.UUID(str(uuid))
            queryset = CampaignApplication.objects.filter(campaign__uuid=uuid)
        except (ValueError):
             # Fallback to ID
            queryset = CampaignApplication.objects.filter(campaign_id=uuid)
            
        apps = queryset.select_related('user', 'user__profile')
        
        count = 1
        for app in apps:
            # Get TikTok data
            tiktok = SocialPlatformAccount.objects.filter(user=app.user, platform='tiktok').first()
            
            tiktok_name = self.clean_text(tiktok.username if tiktok else (app.user.display_name or '-'))
            tiktok_link = tiktok.profile_url if tiktok else '-'
            follower = tiktok.followers_count if tiktok else 0
            
            # Brand Approved Logic (If user is approved or working)
            is_approved = app.status in ['APPROVED', 'WORK_IN_PROGRESS', 'SCRIPT_APPROVED', 
                                       'DRAFT_APPROVED', 'FINAL_APPROVED', 'COMPLETED', 
                                       'PAYMENT_TRANSFERRED']
            brand_approved = 'Yes' if is_approved else ''
            
            # Rate Card info (if visible in profile)
            rate_info = []
            if hasattr(app.user, 'profile'):
                p = app.user.profile
                if p.allow_boost and p.boost_price:
                    rate_info.append(f"Boost: {p.boost_price}")
                if p.allow_original_file and p.original_file_price:
                    rate_info.append(f"Original: {p.original_file_price}")
            rate_str = ", ".join(rate_info)
            
            writer.writerow([
                app.id,
                count,
                tiktok_name,
                tiktok_link,
                follower,
                brand_approved,
                rate_str, # Rate Card
                '', # Condition
                '', # Action
                ''  # Note
            ])
            count += 1
            
        return response


class AdminImportParticipantsView(APIView):
    """
    Import CSV to bulk update statuses.
    """
    permission_classes = [permissions.IsAdminUser]
    
    def post(self, request, uuid):
        import csv
        import io
        
        file = request.FILES.get('file')
        if not file:
            return Response({'error': 'No file uploaded'}, status=400)
            
        try:
            decoded_file = file.read().decode('utf-8-sig').splitlines()
            reader = csv.DictReader(decoded_file)
            
            updated_count = 0
            errors = []
            
            import uuid as uuid_lib
            campaign_id = uuid
            try:
                uuid_obj = uuid_lib.UUID(str(uuid))
                # If valid UUID, resolve to actual numeric ID for foreign key usage if needed, 
                # OR better yet, find the campaign object first to get its ID
                campaign = Campaign.objects.get(uuid=uuid)
                campaign_id = campaign.id
            except (ValueError, Campaign.DoesNotExist):
                # Fallback: assume uuid passed is actually an ID
                campaign_id = uuid

            for row in reader:
                app_id = row.get('Application ID')
                action = row.get('Action (APPROVE/REJECT)', '').upper().strip()
                
                if not app_id or not action:
                    continue
                    
                try:
                    app = CampaignApplication.objects.get(id=app_id, campaign_id=campaign_id)
                    
                    if action == 'APPROVE':
                        # Approve current stage
                        stage = app.current_stage
                        app.approve_stage(stage, feedback="Approved via Bulk Import")
                        updated_count += 1
                        
                    elif action == 'REJECT':
                        stage = app.current_stage
                        app.request_revision(stage, feedback="Rejected via Bulk Import")
                        updated_count += 1
                        
                except CampaignApplication.DoesNotExist:
                   errors.append(f"Application ID {app_id} not found")
                   
            return Response({
                'success': True,
                'updated': updated_count,
                'errors': errors
            })
            
        except Exception as e:
            return Response({'error': str(e)}, status=400)


class AdminCampaignShareLinkView(APIView):
    """
    Generate or retrieve a shareable link for clients (brands).
    """
    permission_classes = [permissions.IsAdminUser]
    
    def post(self, request, uuid):
        try:
            # Try UUID first
            import uuid as uuid_lib
            try:
                uuid_obj = uuid_lib.UUID(str(uuid))
                campaign = Campaign.objects.get(uuid=uuid)
            except (ValueError, Campaign.DoesNotExist):
                # Fallback to ID
                campaign = Campaign.objects.get(id=uuid)
            # Token is auto-generated on save if not exists
            if not campaign.share_token:
                import uuid
                campaign.share_token = str(uuid.uuid4())
                campaign.save()
            
            # Construct the share URL (this should point to a public client-facing page)
            # For now, we'll assume a frontend route exists at /shared/campaign/:token
            base_url = request.build_absolute_uri('/')[:-1]
            # Replace backend port/host with frontend if needed, but for simplicity:
            share_url = f"{settings.CORS_ALLOWED_ORIGINS[0]}/shared/campaign/{campaign.share_token}"
            
            return Response({
                'success': True,
                'shareUrl': share_url,
                'token': campaign.share_token
            })
        except Campaign.DoesNotExist:
            return Response({'error': 'not_found'}, status=404)


class SharedCampaignView(APIView):
    """
    Public endpoint for clients/brands to view campaign progress.
    No authentication required - uses share_token for access.
    """
    permission_classes = [AllowAny]
    
    def get(self, request, token):
        try:
            campaign = Campaign.objects.get(share_token=token)
        except Campaign.DoesNotExist:
            return Response({'error': 'not_found', 'message': 'Campaign not found or link expired'}, status=404)
        
        # Get all applications with user info
        applications = CampaignApplication.objects.filter(campaign=campaign).select_related('user', 'user__profile')
        
        participants = []
        for app in applications:
            # Get TikTok account for display
            tiktok = app.user.social_accounts.filter(platform='tiktok').first()
            
            # Clean display name
            display_name = app.user.display_name or ''
            import re
            display_name = re.sub(r'\?{2,}', '', display_name).strip('?').strip()
            
            # --- Work Links Logic ---
            submission_data = app.submission_data or {}
            
            # Get Script Link (Latest round)
            script_link = None
            if 'script' in submission_data and submission_data['script']:
                scripts = submission_data['script']
                if isinstance(scripts, list) and scripts:
                    script_link = scripts[-1].get('link')
                elif isinstance(scripts, dict):
                    script_link = scripts.get('link')

            # Get Draft Link (Latest round)
            draft_link = None
            if 'draft' in submission_data and submission_data['draft']:
                # Handle List (Multi-Revision) or Dict
                drafts = submission_data['draft']
                if isinstance(drafts, list) and drafts:
                    draft_link = drafts[-1].get('link')
                elif isinstance(drafts, dict):
                    draft_link = drafts.get('link')

            # Get Final Link (Latest round)
            final_link = None
            if 'final' in submission_data and submission_data['final']:
                finals = submission_data['final']
                if isinstance(finals, list) and finals:
                    final_link = finals[-1].get('link')
                elif isinstance(finals, dict):
                    final_link = finals.get('link')

            # --- Payment Slip Logic ---
            # Check query param first, then fall back to campaign setting
            show_slips_param = request.query_params.get('slips', '0')
            show_slips = show_slips_param == '1'
            
            slip_url = None
            if show_slips and app.payment_slip:
                 # Build absolute URL for image
                 slip_url = request.build_absolute_uri(app.payment_slip.url)

            participants.append({
                'id': app.id,
                'display_name': display_name,
                'picture_url': app.user.picture_url,
                'status': app.status,
                'current_stage': app.current_stage,
                'tiktok': {
                    'username': tiktok.username if tiktok else None,
                    'followers': tiktok.followers_count if tiktok else 0,
                    'profile_url': tiktok.profile_url if tiktok else None
                } if tiktok else None,
                'work_links': {
                    'script': script_link,
                    'draft': draft_link,
                    'final': final_link
                },
                'payment_slip_url': slip_url,
                'applied_at': app.applied_at,
                'updated_at': app.updated_at,
                # Insight Data
                'insight_files': app.insight_files,
                'insight_image': app.insight_image,
                'insight_submitted_at': app.insight_submitted_at,
                'insight_data': app.insight_data,
            })
        
        # Summary stats
        stats = {
            'total': len(participants),
            'approved': len([p for p in participants if p['status'] in ['APPROVED', 'WORK_IN_PROGRESS', 'SCRIPT_APPROVED', 'DRAFT_APPROVED', 'FINAL_APPROVED']]),
            'pending_review': len([p for p in participants if p['status'].startswith('SUBMITTED')]),
            'completed': len([p for p in participants if p['status'] in ['COMPLETED', 'PAYMENT_TRANSFERRED']]),
            'waiting': len([p for p in participants if p['status'] == 'WAITING']),
        }
        
        return Response({
            'campaign': {
                'id': campaign.id,
                'title': campaign.title,
                'brand_name': campaign.brand_name,
                'status': campaign.status,
                'show_slip': campaign.show_slip_to_client, # Pass config to frontend
                # Budget removed
                'application_deadline': campaign.application_deadline,
                'content_deadline': campaign.content_deadline,
            },
            'stats': stats,
            'participants': participants
        })


class SharedCampaignExportView(APIView):
    """
    Excel export endpoint for public shared campaign.
    Same data as SharedCampaignView.
    """
    permission_classes = [AllowAny]

    def get(self, request, token):
        try:
            campaign = Campaign.objects.get(share_token=token)
        except Campaign.DoesNotExist:
            return Response({'error': 'not_found', 'message': 'Campaign not found'}, status=404)

        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill
        from django.http import HttpResponse

        # Create Workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Campaign Report"

        # Headers
        headers = [
            'ID', 'Influencer', 'TikTok', 'Followers', 
            'Status' if not campaign.show_slip_to_client else 'Payment Slip', 
            'Script Link', 'Draft Link', 'Final Link', 
            'Last Update'
        ]

        # Style Headers
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="8B5CF6", end_color="8B5CF6", fill_type="solid") # Purple
            cell.alignment = Alignment(horizontal="center")
            
            # Adjust column width
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = 20

        ws.column_dimensions['B'].width = 30 # Name
        ws.column_dimensions['F'].width = 40 # Script
        ws.column_dimensions['G'].width = 40 # Draft
        ws.column_dimensions['H'].width = 40 # Final

        # Data Rows
        applications = CampaignApplication.objects.filter(campaign=campaign).select_related('user')
        
        for row_num, app in enumerate(applications, 2):
            tiktok = app.user.social_accounts.filter(platform='tiktok').first()
            display_name = app.user.display_name or ''
            import re
            display_name = re.sub(r'\?{2,}', '', display_name).strip('?').strip()

            # Links
            submission_data = app.submission_data or {}
            script_link = ''
            if 'script' in submission_data and submission_data['script']:
                data = submission_data['script']
                script_link = (data[-1].get('link') if isinstance(data, list) else data.get('link')) or ''

            draft_link = ''
            if 'draft' in submission_data and submission_data['draft']:
                data = submission_data['draft']
                draft_link = (data[-1].get('link') if isinstance(data, list) else data.get('link')) or ''

            final_link = ''
            if 'final' in submission_data and submission_data['final']:
                data = submission_data['final']
                final_link = (data[-1].get('link') if isinstance(data, list) else data.get('link')) or ''

            # Status or Slip
            status_col_value = app.status
            if campaign.show_slip_to_client:
                 # Logic for Slip: Content is empty (placeholder), Image is added separately
                 status_col_value = ''

            # Base Row
            row = [
                app.id,
                display_name,
                tiktok.username if tiktok else '-',
                tiktok.followers_count if tiktok else 0,
                status_col_value,
                script_link,
                draft_link,
                final_link,
                app.updated_at.strftime('%Y-%m-%d %H:%M')
            ]

            # Write Row
            for col_num, value in enumerate(row, 1):
                cell = ws.cell(row=row_num, column=col_num)
                cell.value = value
                cell.alignment = Alignment(vertical="center")

            # Handle Image Embedding for Column 5 (Payment Slip)
            if campaign.show_slip_to_client and app.payment_slip:
                try:
                    from openpyxl.drawing.image import Image as XLImage
                    from io import BytesIO
                    import requests

                    # Get image URL and download
                    img_url = request.build_absolute_uri(app.payment_slip.url)
                    res = requests.get(img_url, timeout=5)
                    
                    if res.status_code == 200:
                        img_data = BytesIO(res.content)
                        img = XLImage(img_data)
                        
                        # Resize image (Height 80px approx ~ 60pt)
                        # Keep aspect ratio
                        ratio = img.width / img.height
                        new_height = 80
                        new_width = int(new_height * ratio)
                        
                        img.height = new_height
                        img.width = new_width
                        
                        # Anchor to cell (Column 5 is 'E')
                        img.anchor = f'E{row_num}'
                        ws.add_image(img)
                        
                        # Adjust row height to fit image
                        ws.row_dimensions[row_num].height = 65 
                except Exception as e:
                    print(f"Error embedding image for app {app.id}: {e}")
                    # Fallback to URL if image fails? Or just ignore as per 'blank' rule
                    # ws.cell(row=row_num, column=5).value = "Error loading image"
                    pass

        # Response
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename=Campaign_Report_{campaign.id}.xlsx'
        
        # Save invalidates the file pointer if we use BytesIO directly usually, but here we write to response
        wb.save(response)
        return response


class AdminCampaignCreateView(CreateAPIView):
    """
    Create a new campaign with brand logo upload (Admin only).
    """
    permission_classes = [permissions.IsAdminUser]
    serializer_class = CampaignCreateSerializer
    parser_classes = (MultiPartParser, FormParser) # Enable file upload parsing
    
    def perform_create(self, serializer):
        campaign = serializer.save()
        log_action(
            actor=self.request.user,
            action_type='create_campaign',
            target=campaign,
            details={'title': campaign.title, 'brand': campaign.brand_name},
            ip_address=self.request.META.get('REMOTE_ADDR')
        )


class BulkApplicationActionView(APIView):
    """
    Bulk approve or reject campaign applications.
    Expects: { "application_ids": [1, 2, 3], "action": "APPROVE" | "REJECT", "reason": "optional" }
    """
    permission_classes = [permissions.IsAdminUser]

    @transaction.atomic
    def post(self, request):
        from apps.users.services import LineMessagingService
        
        app_ids = request.data.get('application_ids', [])
        action = request.data.get('action')
        stage = request.data.get('stage')  # Optional: script, draft, final, insight
        feedback = request.data.get('feedback', '')

        if not app_ids or not action:
            return Response(
                {'error': 'invalid_data', 'message': 'application_ids and action are required'},
                status=400
            )

        applications = CampaignApplication.objects.filter(id__in=app_ids)
        if not applications.exists():
            return Response({'error': 'not_found', 'message': 'No applications found'}, status=404)

        updated_count = 0
        
        for app in applications:
            # Type 1: Review Stage Work (If stage is provided)
            if stage:
                try:
                    if action == 'APPROVE':
                        app.approve_stage(stage, feedback)
                        updated_count += 1
                    elif action == 'REJECT':
                        app.request_revision(stage, feedback)
                        updated_count += 1
                except Exception as e:
                    print(f"[BulkAction] Skip app {app.id}: {e}")
                    continue
            
            # Type 2: Registration Approval (Legacy/Simple)
            else:
                if action == 'APPROVE':
                    if app.status == 'WAITING':
                        app.status = 'APPROVED'
                        app.save()
                        updated_count += 1
                elif action == 'REJECT':
                    if app.status != 'REJECTED':
                        app.status = 'REJECTED'
                        app.admin_notes = feedback or app.admin_notes
                        app.save()
                        updated_count += 1
        
        return Response({
            'success': True,
            'message': f'Successfully updated {updated_count} applications',
            'updated_count': updated_count
        })

class GenerateFollowUpView(APIView):
    """
    Generate bulk follow-up messages using Gemini.
    """
    permission_classes = [permissions.IsAdminUser]

    def post(self, request):
        try:
            from .services.gemini import GeminiService
            
            users = request.data.get('users', [])
            campaign_title = request.data.get('campaign_title')
            
            results = []
            for user_data in users:
                uid = user_data.get('id')
                name = user_data.get('name')
                status = user_data.get('status')
                due_date = user_data.get('due_date')
                
                message = GeminiService.generate_follow_up_message(
                    influencer_name=name,
                    campaign_title=campaign_title,
                    status=status,
                    deadline=due_date
                )
                
                results.append({
                    'id': uid,
                    'name': name,
                    'status': status,
                    'message': message
                })
                
            return Response({'results': results})
        except Exception as e:
            return Response({'error': str(e), 'traceback': 'Error in GenerateFollowUpView'}, status=500)


class BulkSendMessageView(APIView):
    """
    Send bulk messages via LINE.
    """
    permission_classes = [permissions.IsAdminUser]

    def post(self, request):
        from apps.users.services import LineMessagingService
        from apps.users.models import User
        
        messages_data = request.data.get('messages', [])
        results = []
        sent_count = 0
        failed_count = 0
        
        for item in messages_data:
            passed_id = item.get('id')
            message = item.get('message')
            user_name = item.get('name', 'Unknown')
            
            try:
                # The ID passed from AdminCampaignProject list is the CampaignApplication.id
                # Attempt to get user via application first
                app = CampaignApplication.objects.filter(id=passed_id).select_related('user').first()
                if app:
                    user = app.user
                else:
                    # Fallback to direct User ID if application not found
                    user = User.objects.get(id=passed_id)

                if user.line_user_id:
                    res = LineMessagingService.send_text_message(user.line_user_id, message)
                    if res.get('success'):
                        sent_count += 1
                        results.append({'id': user_id, 'name': user_name, 'status': 'sent'})
                    else:
                        failed_count += 1
                        results.append({'id': user_id, 'name': user_name, 'status': 'failed', 'error': res.get('error', 'LINE API Error')})
                else:
                    failed_count += 1
                    results.append({'id': user_id, 'name': user_name, 'status': 'failed', 'error': 'No LINE ID'})
            except User.DoesNotExist:
                failed_count += 1
                results.append({'id': user_id, 'name': user_name, 'status': 'failed', 'error': 'User not found'})
            except Exception as e:
                failed_count += 1
                results.append({'id': user_id, 'name': user_name, 'status': 'failed', 'error': str(e)})
                
        return Response({
            'success': True,
            'sent': sent_count,
            'failed': failed_count,
            'details': results
        })


class AdminUpdateInfluencerProfileView(APIView):
    """
    PATCH /api/v1/admin/influencers/<pk>/update/
    Allow admin to update influencer profile data (OCR correction).
    """
    permission_classes = [permissions.IsAdminUser]
    
    def patch(self, request, pk):
        try:
            from  apps.users.models import User
            user = User.objects.get(pk=pk)
            # Check if user has profile
            if not hasattr(user, 'profile'):
                 # It might be possible to create one, but for now error
                return Response({'error': 'Profile not found'}, status=404)
            
            profile = user.profile
            data = request.data
            
            # Fields allowed to edit
            editable_fields = [
                'identification_number', 'bank_name', 'bank_account_number',
                'full_name_th', 'house_no', 'village', 'moo', 'soi', 'road',
                'sub_district', 'district', 'province', 'zipcode', 'phone'
            ]
            
            updated_fields = []
            for field in editable_fields:
                if field in data:
                    setattr(profile, field, data[field])
                    updated_fields.append(field)
            
            if updated_fields:
                profile.save()
                
                # Log action if log_action is available
                try:
                    from apps.audit_logs.utils import log_action
                    log_action(
                        actor=request.user,
                        action_type='UPDATE',
                        target=profile, 
                        details={'updated_fields': updated_fields}
                    )
                except ImportError:
                    pass
            
            return Response({'success': True, 'message': 'Updated successfully'})
            
        except User.DoesNotExist:
            return Response({'error': 'User not found'}, status=404)
        except Exception as e:
            logger.error(f"Update Profile Error: {e}")
            return Response({'error': str(e)}, status=500)


class SocialInfoFetchView(APIView):
    """
    Helper endpoint to fetch social media profile info (e.g. YouTube channel)
    without saving it to the database. Used during registration.
    """
    permission_classes = [AllowAny]

    def get(self, request):
        platform = request.query_params.get('platform')
        url = request.query_params.get('url')

        if not platform or not url:
            return Response(
                {'success': False, 'error': 'Platform and URL are required'},
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            # 1. Fetch Profile Data
            result = SocialPlatformService.fetch_profile(platform, url)
            
            if result.get('success'):
                return Response(result)
            else:
                return Response(
                    result,
                    status=status.HTTP_400_BAD_REQUEST
                )

        except Exception as e:
            return Response(
                {'success': False, 'error': str(e)},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

